from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List
from datetime import date
from decimal import Decimal
import io
import zipfile
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image 

# =========================================================
# 1. CARGA DE BASE DE DATOS (Se carga al iniciar el servidor)
# =========================================================
def cargar_base_tiendas(ruta_archivo):
    try:
        df = pd.read_excel(ruta_archivo, dtype=str)
        return df.set_index('TDA').to_dict('index')
    except Exception as e:
        print(f"❌ Error tiendas: {e}")
        return {}

def cargar_tipo_gastos(ruta_archivo):
    try:
        df = pd.read_excel(ruta_archivo, dtype=str)
        return df.set_index('CODIGO').to_dict('index')
    except Exception as e:
        print(f"❌ Error gastos: {e}")
        return {}

# Cargamos en memoria RAM del servidor al iniciar
diccionario_tiendas = cargar_base_tiendas("Copia de BASE DE TIENDAS.xlsx")
diccionario_gastos = cargar_tipo_gastos("TiposGastos.xlsx")

# =========================================================
# 2. ESQUEMAS DE VALIDACIÓN (Lo que enviará el Frontend)
# =========================================================
class GastoItem(BaseModel):
    uidd: str
    cuenta: str
    monto_base: Decimal
    porcentaje_iva: Decimal

class SolicitudPoliza(BaseModel):
    numero_tienda: str
    fecha_poliza: date       # FastAPI espera "YYYY-MM-DD"
    inicio_caja: date
    fin_caja: date
    inicio_ant: date
    fin_ant: date
    cantidad_reembolsada: Decimal
    gastos: List[GastoItem]  # Lista de los gastos capturados

# =========================================================
# 3. INICIO DE LA API
# =========================================================
app = FastAPI(title="Generador de Pólizas SAP")

@app.post("/generar-polizas/")
def generar_polizas(datos: SolicitudPoliza):
    # A) Validar tienda
    if datos.numero_tienda not in diccionario_tiendas:
        raise HTTPException(status_code=404, detail=f"La tienda {datos.numero_tienda} no existe.")
    
    tienda_info = diccionario_tiendas[datos.numero_tienda]
    nombre_tienda = tienda_info.get('PLAZA', '')
    gerente = tienda_info.get('NOMBRE_GERENTE', '')
    cuenta_tienda = tienda_info.get('CUENTA', '')
    fondo = tienda_info.get('CAJA_CHICA', '')
    responsable = tienda_info.get('RESPONSABLE', '')
    supervisor = tienda_info.get('SUPERVISOR', '')

    # B) Procesar fechas y cálculos básicos
    diff_caja = (datos.fin_caja - datos.inicio_caja).days
    diff_ant = (datos.fin_ant - datos.inicio_ant).days

    if diff_caja < 0 or diff_ant < 0:
        raise HTTPException(status_code=400, detail="La fecha de fin no puede ser anterior a la de inicio.")

    # Formatos de texto para Excel
    fecha_poliza_sap = datos.fecha_poliza.strftime("%d.%m.%Y")
    str_inicio_caja = datos.inicio_caja.strftime("%d/%m/%Y")
    str_fin_caja = datos.fin_caja.strftime("%d/%m/%Y")
    str_inicio_ant = datos.inicio_ant.strftime("%d/%m/%Y")
    str_fin_ant = datos.fin_ant.strftime("%d/%m/%Y")

    # C) Procesamiento de Gastos (Poliza Detallada y Agrupada)
    mapa_indices_iva = {Decimal("0.0"): "W0", Decimal("16.0"): "W1", Decimal("8.0"): "W6"}
    poliza_detallada = []
    diccionario_agrupador = {}
    total_gran_factura = Decimal("0")

    for gasto in datos.gastos:
        if gasto.cuenta not in diccionario_gastos:
            raise HTTPException(status_code=400, detail=f"Cuenta {gasto.cuenta} no válida en catálogo.")
        
        desc = diccionario_gastos[gasto.cuenta]
        gasto_descripcion = list(desc.values())[0] if isinstance(desc, dict) else desc

        iva_calculado = (gasto.monto_base / (1 + gasto.porcentaje_iva / 100)) * (gasto.porcentaje_iva / 100)
        subtotal_factura = gasto.monto_base - iva_calculado
        indice_iva = mapa_indices_iva.get(gasto.porcentaje_iva, "W_ND")

        registro = {
            "UIDD": gasto.uidd,
            "Cuenta": gasto.cuenta,
            "Identificador": "S",
            "Descripcion": gasto_descripcion,
            "Indice_IVA": indice_iva,
            "IVA": iva_calculado,
            "Subtotal": subtotal_factura,
            "Total": gasto.monto_base,
            "Tienda": datos.numero_tienda
        }
        poliza_detallada.append(registro)
        total_gran_factura += gasto.monto_base

        # Agrupador
        if gasto.cuenta in diccionario_agrupador:
            diccionario_agrupador[gasto.cuenta]['Total'] += gasto.monto_base
            diccionario_agrupador[gasto.cuenta]['Subtotal'] += subtotal_factura
            diccionario_agrupador[gasto.cuenta]['IVA'] += iva_calculado
            diccionario_agrupador[gasto.cuenta]['Cantidad_facturas'] += 1
        else:
            nuevo = registro.copy()
            nuevo['Cantidad_facturas'] = 1
            nuevo['UIDD'] = "Varios"
            diccionario_agrupador[gasto.cuenta] = nuevo

    poliza_agrupada = list(diccionario_agrupador.values())

    # La 'K' (Acreedor)
    registro_acreedor = {
        "UIDD": "Varios", "Cuenta": "Acreedores Diversos / Proveedor", "Identificador": "K",
        "Descripcion": "Total", "Indice_IVA": "N/A", "IVA": Decimal("0"), "Subtotal": Decimal("0"),
        "Total": -total_gran_factura, "Tienda": datos.numero_tienda, "Cantidad_facturas": 0
    }
    poliza_detallada.append(registro_acreedor)
    poliza_agrupada.append(registro_acreedor)

    # D) Creación del Archivo 1: SAP (En memoria)
    wb_sap = Workbook()
    ws_sap = wb_sap.active
    ws_sap.title = "Datos"


    # Tamaño de celdas
    ws_sap.column_dimensions['A'].width = 8.0
    ws_sap.column_dimensions['B'].width = 8.0
    ws_sap.column_dimensions['C'].width = 9.1
    ws_sap.column_dimensions['D'].width = 9.1
    ws_sap.column_dimensions['E'].width = 8.0
    ws_sap.column_dimensions['F'].width = 15.0
    ws_sap.column_dimensions['G'].width = 47.0
    ws_sap.column_dimensions['H'].width = 47.0
    
    # Formato SAP...
    color_fondo = PatternFill(start_color="FFCDCD", end_color="FFCDCD", fill_type="solid")
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_sap[f"{col}1"].fill = color_fondo

    # Insertar datos en celdas
    ws_sap['A1'] = 'IAC1'# En esta columna va el identificador del gasto
    ws_sap['B1'] = 'KR' # Si el identificador es K, entonces va el número de tienda
    ws_sap['C1'] = f'{fecha_poliza_sap}' # Si el identificador es K, entonces va el total de la póliza (en negativo)
    ws_sap['D1'] = f'{fecha_poliza_sap}' # Si el identificador es K, no se escribe nada
    ws_sap['E1'] = f'MXN' # En esta columna va el número de tienda, se repita por cada S que haya
    ws_sap['F1'] = f'{datos.numero_tienda} CAJA CHICA' 
    ws_sap['G1'] = f'{datos.numero_tienda} {str_inicio_caja} AL {str_fin_caja} {gerente}' 
    ws_sap['H1'] = f' ' # Si el identificador es K, se escribe {tienda} {inicio_caja} AL {fin_caja} 

    for mov in poliza_detallada:
        es_k = (mov['Identificador'] == 'K')
        col_b = datos.numero_tienda if es_k else mov['Cuenta']
        col_d = '' if es_k else mov['Indice_IVA']
        col_e = '' if es_k else datos.numero_tienda
        col_h = f'{datos.numero_tienda} {str_inicio_caja} AL {str_fin_caja} {gerente}' if es_k else mov['UIDD']
        ws_sap.append([mov['Identificador'], col_b, float(mov['Total']), col_d, col_e, '', f'{datos.numero_tienda} CAJA CHICA', col_h])

    # E) Creación del Archivo 2: Solicitud (En memoria)
    wb_solicitud = load_workbook('COPIA FORMATO REEMBOLSO.xlsx')
    ws_solicitud = wb_solicitud.active
    
    try:
        logo = Image('logoV.png')
        ws_solicitud.add_image(logo, 'C3')
    except Exception as e:
        print("Logo no insertado:", e)

    ws_solicitud['I9'] = fecha_poliza_sap
    ws_solicitud['I11'] = f'{datos.numero_tienda} - {nombre_tienda}'
    ws_solicitud['D14'] = gerente
    ws_solicitud['D16'] = cuenta_tienda
    ws_solicitud['D18'] = float(total_gran_factura)
    ws_solicitud['D20'] = fondo
    ws_solicitud['F29'] = str_inicio_caja
    ws_solicitud['I29'] = str_fin_caja
    ws_solicitud['F31'] = str_inicio_ant
    ws_solicitud['I31'] = str_fin_ant
    ws_solicitud['H34'] = responsable
    ws_solicitud['K34'] = supervisor
    ws_solicitud['L29'] = f"{diff_caja} días"
    ws_solicitud['L31'] = f"{diff_ant} días"
    ws_solicitud['K31'] = float(-datos.cantidad_reembolsada)
    
    suma_ivas = sum(item['IVA'] for item in poliza_detallada if item['Identificador'] == 'S')
    ws_solicitud['I65'] = float(total_gran_factura)
    ws_solicitud['I68'] = float(total_gran_factura)
    ws_solicitud['G63'] = float(suma_ivas)

    fila_actual = 41
    for mov in poliza_agrupada:
        if mov['Identificador'] == 'S':
            ws_solicitud[f'C{fila_actual}'] = mov['Cuenta']
            ws_solicitud[f'D{fila_actual}'] = mov['Descripcion']
            ws_solicitud[f'G{fila_actual}'] = float(mov['Subtotal'])
            ws_solicitud[f'E{fila_actual}'] = mov['Cantidad_facturas']
            fila_actual += 1

    # F) Guardar ambos en RAM y Empaquetar en un ZIP
    buffer_sap = io.BytesIO()
    wb_sap.save(buffer_sap)
    
    buffer_solicitud = io.BytesIO()
    wb_solicitud.save(buffer_solicitud)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        # Aquí definimos cómo se llamarán los archivos DENTRO del ZIP
        zip_file.writestr(f"{datos.numero_tienda} CAJA CHICA.xlsx", buffer_sap.getvalue())
        zip_file.writestr(f"{datos.numero_tienda} Poliza Reembolso.xlsx", buffer_solicitud.getvalue())

    # Reiniciamos el cursor del buffer de ZIP al inicio antes de enviarlo
    zip_buffer.seek(0)

    # Devolvemos el archivo ZIP generado
    return StreamingResponse(
        zip_buffer, 
        media_type="application/x-zip-compressed",
        headers={"Content-Disposition": f"attachment; filename=Polizas_{datos.numero_tienda}.zip"}
    )