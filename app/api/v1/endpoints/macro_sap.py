import io
import os
import re
import uuid
import zipfile
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from typing import Annotated
from zoneinfo import ZoneInfo 

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.db.session import get_db

from app.services.tax_rules import (
    normalizar_texto,
    cargar_tipo_gastos,
    crear_indice_categorias,
    cargar_tiendas_iva_w6,
    determinar_iva_e_indice
)

from app.services.spending_summary import obtener_resumen_gasto_tienda
# 1. Definir el Router en lugar de la App
router = APIRouter()

# 2. Configurar las rutas absolutas para leer tus archivos (para que no falle al ejecutarlo)
# Esto calcula la ruta basándose en dónde está este archivo macro_sap.py
BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))
ASSETS_DIR = os.path.join(BASE_DIR, "assets")

ruta_tiendas = os.path.join(ASSETS_DIR, "Copia de BASE DE TIENDAS.xlsx")
ruta_gastos = os.path.join(ASSETS_DIR, "TiposGastos.xlsx")
ruta_plantilla = os.path.join(ASSETS_DIR, "COPIA FORMATO REEMBOLSO.xlsx")
ruta_logo = os.path.join(ASSETS_DIR, "logoV.png")
ruta_W6 = os.path.join(ASSETS_DIR, "TDAS IVA W6.xlsx")
ruta_saldos = os.path.join(ASSETS_DIR, "COPIA SALDO DE TDAS ANOS 24 25 Y 26.xlsx")


def limpiar_nombre_archivo(valor: str) -> str:
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", valor).strip()


def cargar_base_tiendas(ruta_archivo):
    try:
        df = pd.read_excel(ruta_archivo, dtype=str).fillna("")

        df["TDA"] = df["TDA"].str.strip()

        return df.set_index("TDA").to_dict("index")

    except Exception as e:  # noqa: BLE001
        print(f"❌ Error tiendas: {e}")
        return {}


def convertir_decimal(valor):
    try:
        texto = str(valor).strip()

        if not texto or texto.lower() == "nan":
            return Decimal("0.00")

        texto = texto.replace("\u00A0", "")
        texto = texto.replace("$", "")
        texto = texto.replace(",", "")

        if texto.startswith("(") and texto.endswith(")"):
            texto = f"-{texto[1:-1]}"

        return Decimal(texto).quantize(
            Decimal("0.01")
        )

    except (InvalidOperation, ValueError) as exc:
        raise ValueError(
            f"Saldo no numérico encontrado: {valor!r}"
        ) from exc


def cargar_saldos_tiendas(ruta_archivo):
    try:
        df = pd.read_excel(
            ruta_archivo,
            sheet_name="PRESUPUESTO 2024-2026",
            usecols="A,D,E",
            header=None,
            dtype=str,
        ).fillna("")

        # Asignamos nombres internos sin depender de los encabezados visuales
        df.columns = [
            "TDA",
            "GASTO_2025",
            "GASTO_2026",
        ]

        # Limpiar texto en la columna de tienda
        df["TDA"] = df["TDA"].astype(str).str.strip()

        # Solo aceptamos códigos reales de tienda:
        # V001, V042, V999, etc.
        patron_tienda = patron_tienda = r"^[A-Z]+\d+$"

        df = df[
            df["TDA"].str.fullmatch(
                patron_tienda,
                case=False,
                na=False,
            )
        ].copy()

        # Normalizar el código final como mayúsculas
        df["TDA"] = df["TDA"].str.upper()

        if df.empty:
            raise ValueError(
                "No se encontraron tiendas con formato V###. "
                "Revisa la columna A y el patrón de los códigos."
            )

        # Convertir únicamente datos de filas reales de tienda
        df["GASTO_2025"] = df["GASTO_2025"].apply(
            convertir_decimal
        )

        df["GASTO_2026"] = df["GASTO_2026"].apply(
            convertir_decimal
        )

        # Validar códigos duplicados
        if df["TDA"].duplicated().any():
            duplicados = df.loc[
                df["TDA"].duplicated(keep=False),
                "TDA",
            ].tolist()

            raise ValueError(
                f"Hay tiendas duplicadas en el archivo: {duplicados}"
            )

        return df.set_index("TDA").to_dict("index")

    except Exception as e:
        print(f"❌ Error al cargar saldos de tiendas: {e}")
        return {}


# Cargamos en memoria RAM del servidor al iniciar
diccionario_tiendas = cargar_base_tiendas(ruta_tiendas)
diccionario_gastos = cargar_tipo_gastos(ruta_gastos)
indice_categorias = crear_indice_categorias(diccionario_gastos)
diccionario_saldos = cargar_saldos_tiendas(ruta_saldos)
tiendas_iva_w6 = cargar_tiendas_iva_w6(ruta_W6)

# =========================================================
# 3. INICIO DE LA API
# =========================================================

@router.post("/generar-polizas/{solicitud_id}")
def generar_polizas(
    solicitud_id: uuid.UUID,
    db: Annotated[Session, Depends(get_db)],
):

    # ========================================================
    # 1. OBTENER LA SOLICITUD DESDE LA BASE DE DATOS
    # ========================================================
    # Hacemos una consulta directa (query) a la tabla que me mostraste
    query_solicitud = text("""
        SELECT *
        FROM reimbursement_requests
        WHERE id = :id
    """)

    solicitud = db.execute(
        query_solicitud,
        {"id": solicitud_id},
    ).mappings().first()

    if solicitud is None:
        raise HTTPException(
            status_code=404,
            detail="Solicitud de reembolso no encontrada.",
        )


    # Para obtener el ID de la tienda y el ID del periodo (que usaremos más abajo)
    store_id = solicitud['store_id']
    period_id = solicitud['period_id']
    query_tienda = text("SELECT * FROM stores WHERE id = :id")
    tienda_db = db.execute(query_tienda, {"id": store_id}).mappings().first()

    if tienda_db is None:
        raise HTTPException(
            status_code=404,
            detail=f"No existe la tienda asociada al store_id {store_id}.",
        )

    numero_tienda = str(tienda_db["code"]).strip()

    query_periodo = text("""
        SELECT *
        FROM periods
        WHERE id = :id
    """)

    periodo = db.execute(
        query_periodo,
        {"id": period_id},
    ).mappings().first()

    if periodo is None:
        raise HTTPException(
            status_code=404,
            detail=f"No existe el periodo asociado al period_id {period_id}.",
        )

    # Cambia estos nombres si periods usa nombres distintos.
    inicio_caja = periodo["starts_on"]
    fin_caja = periodo["ends_on"]

    inicio_ant = solicitud["previous_reimbursement_starts_on"]
    fin_ant = solicitud["previous_reimbursement_ends_on"]
    cantidad_reembolsada = Decimal(
        str(solicitud["previous_reimbursement_amount"] or "0")
    )

    # Fechas del periodo actual de caja
    inicio_caja = periodo["starts_on"]
    fin_caja = periodo["ends_on"]

    # Fechas del periodo anterior
    inicio_ant = solicitud["previous_reimbursement_starts_on"]
    fin_ant = solicitud["previous_reimbursement_ends_on"]

    # Fecha de creación de la solicitud
    created_at = solicitud["created_at"]

    # Zona horaria oficial para mostrar fechas
    zona_mexico = ZoneInfo("America/Mexico_City")

    if created_at is not None:
        if created_at.tzinfo is None:
            created_at = created_at.replace(tzinfo=UTC)

        fecha_poliza_obj = created_at.astimezone(zona_mexico).date()
    else:
        fecha_poliza_obj = datetime.now(zona_mexico).date()

    # Validar periodo actual
    if inicio_caja is None or fin_caja is None:
        raise HTTPException(
            status_code=422,
            detail="La solicitud no tiene un periodo de caja completo.",
        )

    diff_caja = (fin_caja - inicio_caja).days

    # Validar periodo anterior únicamente si existe
    if inicio_ant is not None and fin_ant is not None:
        diff_ant = (fin_ant - inicio_ant).days
    else:
        diff_ant = 0

    if diff_caja < 0:
        raise HTTPException(
            status_code=422,
            detail="El fin del periodo de caja no puede ser anterior al inicio.",
        )

    if diff_ant < 0:
        raise HTTPException(
            status_code=422,
            detail="El fin del periodo anterior no puede ser anterior al inicio.",
        )

    # Formatos requeridos por el Excel
    fecha_poliza_sap = fecha_poliza_obj.strftime("%d.%m.%Y")

    str_inicio_caja = inicio_caja.strftime("%d/%m/%Y")
    str_fin_caja = fin_caja.strftime("%d/%m/%Y")

    str_inicio_ant = (
        inicio_ant.strftime("%d/%m/%Y")
        if inicio_ant is not None
        else ""
    )

    str_fin_ant = (
        fin_ant.strftime("%d/%m/%Y")
        if fin_ant is not None
        else ""
    )

    # ========================================================
    # 3. CONECTAR CON TU BASE DE DATOS LOCAL DE EXCEL
    # ========================================================
    # Ya que tenemos el numero_tienda de la base de datos (ej. "V101"), 
    # consultamos tu Excel como siempre lo hemos hecho:
    if numero_tienda not in diccionario_tiendas:
        raise HTTPException(
            status_code=422,
            detail=(
                f"La tienda '{numero_tienda}' existe en SQL, "
                "pero no en Copia de BASE DE TIENDAS.xlsx."
            ),
        )

    tienda_info = diccionario_tiendas[numero_tienda]

    nombre_tienda = tienda_info.get("PLAZA", "")
    gerente = tienda_info.get("NOMBRE_GERENTE", "")
    cuenta_tienda = tienda_info.get("CUENTA", "")
    fondo = tienda_info.get("CAJA_CHICA", "")
    responsable = tienda_info.get("RESPONSABLE", "")
    supervisor = tienda_info.get("SUPERVISOR", "")

    # ========================================================
    # 4. OBTENER LOS GASTOS VINCULADOS A ESTA SOLICITUD
    # ========================================================
    # En la tabla 'expenses', veo que hay una columna 'reimbursement_request_id'
    query_gastos = text("""
        SELECT *
        FROM expenses
        WHERE reimbursement_request_id = :request_id
          AND removed_at IS NULL
        ORDER BY spent_on, created_at, id
    """)

    gastos_db = db.execute(
        query_gastos,
        {"request_id": solicitud_id},
    ).mappings().all()

    if not gastos_db:
        raise HTTPException(
            status_code=422,
            detail="La solicitud no tiene gastos activos asociados.",
        )

    gastos_invalidos = [
        str(gasto["id"])
        for gasto in gastos_db
        if gasto["status"] in {"removed", "rejected"}
    ]

    if gastos_invalidos:
        raise HTTPException(
            status_code=422,
            detail=f"Hay gastos no procesables: {gastos_invalidos}",
        )

    # ========================================================
    # 5. CONSTRUIR LA PÓLIZA CON LA LÓGICA FINANCIERA
    # ========================================================
    poliza_detallada = []
    diccionario_agrupador = {}
    total_gran_factura = Decimal(0)
    porcentaje_iva_default = Decimal(16)

    for gasto in gastos_db:
        categoria_bd = str(gasto["category"] or "").strip()
        clave_categoria = normalizar_texto(categoria_bd)

        cuenta_info = indice_categorias.get(clave_categoria)

        if cuenta_info is None:
            raise HTTPException(
                status_code=422,
                detail=(
                    f"La categoría '{categoria_bd}' del gasto "
                    f"{gasto['id']} no tiene equivalencia en "
                    "TiposGastos.xlsx."
                ),
            )

        cuenta_gasto = cuenta_info["codigo"]
        gasto_descripcion = cuenta_info["descripcion"]

        uidd_factura = (
            str(gasto["cfdi_uuid"]).strip()
            if gasto["cfdi_uuid"]
            else "Sin Folio"
        )

        monto_total = Decimal(str(gasto["amount"]))

        tasa_iva_bd = (
            Decimal(str(gasto["cfdi_tax_rate"]))
            if gasto["cfdi_tax_rate"] is not None
            else porcentaje_iva_default
        )

        porcentaje_iva, indice_iva = determinar_iva_e_indice(
            descripcion=gasto_descripcion,
            numero_tienda=numero_tienda,
            porcentaje_iva=tasa_iva_bd,
        )

        if (gasto["cfdi_tax_amount"] is not None and gasto["cfdi_subtotal"] is not None and porcentaje_iva not in {Decimal(0),}):
            iva_calculado = Decimal(
                str(gasto["cfdi_tax_amount"])
            )
            subtotal_factura = Decimal(
                str(gasto["cfdi_subtotal"])
            )
        elif porcentaje_iva == Decimal(0):
            subtotal_factura = monto_total
            iva_calculado = Decimal(0)
        else:
            subtotal_factura = (
                monto_total
                / (
                    Decimal(1)
                    + porcentaje_iva / Decimal(100)
                )
            )
            iva_calculado = monto_total - subtotal_factura


        registro = {
            "UIDD": uidd_factura,
            "Cuenta": cuenta_gasto,
            "Identificador": "S",
            "Descripcion": gasto_descripcion,
            "Porcentaje_IVA": porcentaje_iva,
            "Indice_IVA": indice_iva,
            "IVA": iva_calculado,
            "Subtotal": subtotal_factura,
            "Total": monto_total,
            "Tienda": numero_tienda,
            "Expense_ID": str(gasto["id"]),
        }

        poliza_detallada.append(registro)
        total_gran_factura += monto_total

        if cuenta_gasto in diccionario_agrupador:
            agrupado = diccionario_agrupador[cuenta_gasto]

            agrupado["Total"] += monto_total
            agrupado["Subtotal"] += subtotal_factura
            agrupado["IVA"] += iva_calculado
            agrupado["Cantidad_facturas"] += 1
        else:
            agrupado = registro.copy()
            agrupado["UIDD"] = "Varios"
            agrupado["Cantidad_facturas"] = 1

            diccionario_agrupador[cuenta_gasto] = agrupado

    poliza_agrupada = list(diccionario_agrupador.values())

    registro_acreedor = {
        "UIDD": "Varios",
        "Cuenta": "Acreedores Diversos / Proveedor",
        "Identificador": "K",
        "Descripcion": "Total Cuenta por Pagar",
        "Porcentaje_IVA": Decimal(0),
        "Indice_IVA": "N/A",
        "IVA": Decimal(0),
        "Subtotal": Decimal(0),
        "Total": -total_gran_factura,
        "Tienda": numero_tienda,
        "Cantidad_facturas": 0,
    }

    poliza_detallada.append(registro_acreedor)
    poliza_agrupada.append(registro_acreedor)

    # D) Creación del Archivo 1: SAP (En memoria)
    wb_sap = Workbook()
    ws_sap = wb_sap.active
    ws_sap.title = "Datos"


    # Tamaño de celdas
    ws_sap.column_dimensions['A'].width = 8.0
    ws_sap.column_dimensions['B'].width = 8.0
    ws_sap.column_dimensions['C'].width = 9.1
    ws_sap.column_dimensions['D'].width = 9.1
    ws_sap.column_dimensions['E'].width = 8.0
    ws_sap.column_dimensions['F'].width = 15.0
    ws_sap.column_dimensions['G'].width = 47.0
    ws_sap.column_dimensions['H'].width = 47.0
    
    # Formato SAP...
    color_fondo = PatternFill(start_color="FFCDCD", end_color="FFCDCD", fill_type="solid")
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_sap[f"{col}1"].fill = color_fondo

    # Insertar datos en celdas
    ws_sap['A1'] = 'IAC1'# En esta columna va el identificador del gasto
    ws_sap['B1'] = 'KR' # Si el identificador es K, entonces va el número de tienda
    ws_sap['C1'] = f'{fecha_poliza_sap}' # Si el identificador es K, entonces va el total de la póliza (en negativo)
    ws_sap['D1'] = f'{fecha_poliza_sap}' # Si el identificador es K, no se escribe nada
    ws_sap['E1'] = 'MXN' # En esta columna va el número de tienda, se repita por cada S que haya
    ws_sap['F1'] = f'{numero_tienda} CAJA CHICA' 
    ws_sap['G1'] = f'{numero_tienda} {str_inicio_caja} AL {str_fin_caja} {gerente}' 
    ws_sap['H1'] = ' ' # Si el identificador es K, se escribe {tienda} {inicio_caja} AL {fin_caja}

    for mov in poliza_detallada:
        es_k = (mov['Identificador'] == 'K')

        col_b = numero_tienda if es_k else mov['Cuenta']
        col_d = '' if es_k else mov['Indice_IVA']
        col_e = '' if es_k else numero_tienda

        col_h = f'{numero_tienda} {str_inicio_caja} AL {str_fin_caja} {gerente}' if es_k else mov['UIDD']
        ws_sap.append([mov['Identificador'], col_b, float(mov['Total']), col_d, col_e, "", f'{numero_tienda} CAJA CHICA', col_h])


    # E) Creación del Archivo 2: Solicitud (En memoria)
    wb_solicitud = load_workbook(ruta_plantilla)
    ws_solicitud = wb_solicitud.active

    resumen_2025 = obtener_resumen_gasto_tienda(
        db=db,
        store_id=store_id,
        fiscal_year=2025,
    )

    resumen_2026 = obtener_resumen_gasto_tienda(
        db=db,
        store_id=store_id,
        fiscal_year=2026,
    )
    
    try:
        logo = Image(ruta_logo)
        ws_solicitud.add_image(logo, 'C3')
    except Exception as e:  # noqa: BLE001
        print("Logo no insertado:", e)

    ws_solicitud['I9'] = fecha_poliza_sap
    ws_solicitud['I11'] = f'{numero_tienda} - {nombre_tienda}'
    ws_solicitud['D14'] = gerente
    ws_solicitud['D16'] = cuenta_tienda
    ws_solicitud['D18'] = float(total_gran_factura)
    ws_solicitud['D20'] = fondo
    ws_solicitud['F29'] = str_inicio_caja
    ws_solicitud['I29'] = str_fin_caja
    ws_solicitud['F31'] = str_inicio_ant
    ws_solicitud['I31'] = str_fin_ant
    ws_solicitud['H34'] = responsable
    ws_solicitud['K34'] = supervisor
    ws_solicitud['L29'] = f"{diff_caja} días"
    ws_solicitud['L31'] = f"{diff_ant} días"
    ws_solicitud['K31'] = float(cantidad_reembolsada)
    
    suma_ivas = sum(item['IVA'] for item in poliza_detallada if item['Identificador'] == 'S')
    ws_solicitud['I65'] = float(total_gran_factura)
    ws_solicitud['I68'] = float(total_gran_factura)
    ws_solicitud['G63'] = float(suma_ivas)

    ws_solicitud["E22"] = float(resumen_2025["current_accumulated"])
    ws_solicitud["E24"] = float(resumen_2026["current_accumulated"])
    fila_actual = 41
    
    for mov in poliza_agrupada:
        if mov['Identificador'] == 'S':
            ws_solicitud[f'C{fila_actual}'] = mov['Cuenta']
            ws_solicitud[f'D{fila_actual}'] = mov['Descripcion']
            ws_solicitud[f'G{fila_actual}'] = float(mov['Subtotal'])
            ws_solicitud[f'E{fila_actual}'] = mov['Cantidad_facturas']
            fila_actual += 1

    # F) Guardar ambos en RAM y Empaquetar en un ZIP
    buffer_sap = io.BytesIO()
    wb_sap.save(buffer_sap)
    
    buffer_solicitud = io.BytesIO()
    wb_solicitud.save(buffer_solicitud)

    
    folio = str(solicitud["folio"] or solicitud_id).strip()
    folio_archivo = limpiar_nombre_archivo(folio)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        # Aquí definimos cómo se llamarán los archivos DENTRO del ZIP
        zip_file.writestr(f"CAJA CHICA {folio}.xlsx", buffer_sap.getvalue())
        zip_file.writestr(f"Poliza Reembolso {folio}.xlsx", buffer_solicitud.getvalue())

    # Reiniciamos el cursor del buffer de ZIP al inicio antes de enviarlo
    zip_buffer.seek(0)

    nombre_zip = f"Poliza_{folio_archivo}.zip"

    return StreamingResponse(
        zip_buffer,
        media_type="application/x-zip-compressed",
        headers={
            "Content-Disposition": (f'attachment; filename="{nombre_zip}"')
        },
    )
